#!/usr/bin/env python3
"""Build a formatted Excel workbook from businesses-free.csv.

Layout: a Summary sheet, then one sheet per zip code (sorted ascending),
each listing that zip's businesses with phone / email / website / contact page.
"""
import csv, sys, re
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv) > 1 else "businesses-free.csv"
OUT = sys.argv[2] if len(sys.argv) > 2 else "TKE-StLouis-Businesses.xlsx"

RED = "AE2624"        # TKE burgundy
RED_DK = "7E1A18"
GREY = "F2F2F2"
FONT = "Arial"

EMAIL_FIND = re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
def clean_email(val):
    """Extract the first valid email; drop scraped HTML/JS/template junk."""
    val = (val or "").strip()
    if not val or "{{" in val or "${" in val or "&#" in val:
        return ""
    m = EMAIL_FIND.search(val)
    if not m:
        return ""
    e = m.group(0).rstrip(".").lower()
    return e if 5 <= len(e) <= 80 else ""

with open(SRC, newline="", encoding="utf-8") as f:
    rows = list(csv.DictReader(f))

# Sanitize scraped emails (strip HTML/JS junk, recover valid addresses).
cleaned = 0
for r in rows:
    orig = (r.get("email") or "").strip()
    fixed = clean_email(orig)
    if fixed != orig:
        cleaned += 1
    r["email"] = fixed
if cleaned:
    print(f"Sanitized {cleaned} email cell(s).")
    with open(SRC, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)

# Column spec: (csv_key, header_title, width). event_fit only included if present.
COLSPEC = [
    ("name",         "Business Name", 34),
    ("category",     "Category",      22),
    ("event_fit",    "Event Fit",     12),
    ("address",      "Address",       42),
    ("phone",        "Phone",         18),
    ("email",        "Email",         32),
    ("website",      "Website",       38),
    ("contact_page", "Contact Page",  38),
]
present = set(rows[0].keys()) if rows else set()
COLSPEC = [c for c in COLSPEC if c[0] in present]
headers = [c[0] for c in COLSPEC]
titles  = [c[1] for c in COLSPEC]
widths  = [c[2] for c in COLSPEC]
NCOL = len(headers)
LASTCOL = get_column_letter(NCOL)

# Group rows by zip (preserve a city label per zip).
by_zip = OrderedDict()
city_of = {}
for r in rows:
    z = (r.get("zip") or "").strip()
    if not z:
        continue
    by_zip.setdefault(z, []).append(r)
    city_of.setdefault(z, (r.get("city") or "").strip())

zips_sorted = sorted(by_zip)

# Styles
hdr_fill = PatternFill("solid", fgColor=RED)
hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
cell_font = Font(name=FONT, size=10)
link_font = Font(name=FONT, size=10, color="0563C1", underline="single")
stripe = PatternFill("solid", fgColor=GREY)
border = Border(bottom=Side(style="thin", color="DDDDDD"))

FORBIDDEN = re.compile(r'[:\\/?*\[\]]')
def sheet_name(z):
    city = city_of.get(z, "")
    name = f"{z} {city}".strip()
    name = FORBIDDEN.sub("", name)
    return name[:31]

def fill_sheet(ws, zip_rows):
    for c, t in enumerate(titles, 1):
        cell = ws.cell(1, c, t)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 22
    for ri, row in enumerate(sorted(zip_rows, key=lambda x: (x.get("name") or "").lower()), 2):
        for c, key in enumerate(headers, 1):
            val = (row.get(key) or "").strip()
            cell = ws.cell(ri, c)
            cell.font = cell_font; cell.border = border
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                cell.fill = stripe
            if not val:
                continue
            if key == "email":
                cell.value = val; cell.hyperlink = f"mailto:{val}"; cell.font = link_font
            elif key in ("website", "contact_page"):
                cell.value = val; cell.hyperlink = val; cell.font = link_font
            else:
                cell.value = val
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    last = len(zip_rows) + 1
    ws.auto_filter.ref = f"A1:{LASTCOL}{last}"

wb = Workbook()

# ----------------------------------------------------------------- Summary
s = wb.active
s.title = "Summary"
s.sheet_view.showGridLines = False
for col, w in zip("ABCDE", (10, 22, 10, 12, 12)):
    s.column_dimensions[col].width = w

s.cell(1, 1, "TKE Epsilon Alpha — St. Louis Area Business Database")\
 .font = Font(name=FONT, bold=True, size=15, color=RED_DK)
s.cell(2, 1, "One sheet per zip code · source: OpenStreetMap + website scraping")\
 .font = Font(name=FONT, italic=True, size=9, color="808080")

def cnt(rs, key): return sum(1 for r in rs if (r.get(key) or "").strip())
reach = lambda rs: sum(1 for r in rs if (r.get("email") or "").strip() or (r.get("contact_page") or "").strip())

totals = [
    ("Total businesses", len(rows)),
    ("With phone", cnt(rows, "phone")),
    ("With email", cnt(rows, "email")),
    ("With contact-page link", cnt(rows, "contact_page")),
    ("Reachable (email or contact)", reach(rows)),
    ("Zip codes", len(zips_sorted)),
]
s.cell(4, 1, "Totals").font = Font(name=FONT, bold=True, size=12, color=RED_DK)
for i, (label, val) in enumerate(totals, 5):
    s.cell(i, 1, label).font = Font(name=FONT, size=11)
    c = s.cell(i, 2, val); c.font = Font(name=FONT, bold=True, size=11)
    c.number_format = "#,##0"; c.alignment = Alignment(horizontal="right")

# Per-zip index table with counts (Zip | City | Businesses | Phone | Email)
hdr_row = 12
for c, t in enumerate(["Zip", "City", "Businesses", "Phone", "Email"], 1):
    cell = s.cell(hdr_row, c, t)
    cell.fill = hdr_fill; cell.font = hdr_font
for i, z in enumerate(zips_sorted, hdr_row + 1):
    zr = by_zip[z]
    s.cell(i, 1, z).font = Font(name=FONT, size=10)
    s.cell(i, 2, city_of.get(z, "")).font = Font(name=FONT, size=10)
    for col, key in ((3, None), (4, "phone"), (5, "email")):
        v = len(zr) if key is None else cnt(zr, key)
        cc = s.cell(i, col, v); cc.font = Font(name=FONT, size=10)
        cc.alignment = Alignment(horizontal="right")
s.freeze_panes = f"A{hdr_row + 1}"

# ----------------------------------------------------------------- Per-zip sheets
for z in zips_sorted:
    ws = wb.create_sheet(sheet_name(z))
    fill_sheet(ws, by_zip[z])

wb.save(OUT)
print(f"Wrote {OUT}: Summary + {len(zips_sorted)} zip sheets, {len(rows)} businesses.")
